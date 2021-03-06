import re
import traceback
import json
import argh
import random

from openpyxl import Workbook
from openpyxl import load_workbook

class Pair(object):
    '''
    Represents a pairing in a round.
    If the scores are no present that means the round has not been concluded
    '''

    def __init__(self, player1, player2, score1, score2):
        if player1 == 'Bye' or player2 == 'Bye' or player1 < player2:
            self.player2 = player2
            self.player1 = player1
            self.score2 = score2
            self.score1 = score1
        else :
            self.player1 = player2
            self.player2 = player1
            self.score1 = score2
            self.score2 = score1

    def __eq__(self, other):
        if isinstance(other, Pair):
            if self.player1 == other.player1 and self.player2 == other.player2:
                if self.score1 == other.score1 and self.score2 == other.score2:
                    return True

        False

    def to_array(self):
        return [self.player1, self.score1, self.player2, self.score2]


class Tsh(object):
    '''
    Imports data from TSH
    TSH saves it's data in what are known as division files. Typically in a tournament 
    with only one division that file will be named as `a.t` hence division files are 
    sometimes known as at files as well
    '''

    def __init__(self, filename):
        self.filename = filename
        self.players = [{'name': 'Bye'}]
        self.rounds = 0;


    def create_division(self,players):
        '''
        Creates a division file. 
        The players list is a tuple of the form ('Player name','Rating')
        '''
        for player in players:
            print player[0],",",player[1]


    def process_data(self):
        players = self.players

        with open(self.filename) as f:
            for line in f:
                if line and len(line) > 30 :
                    rating = re.search('[0-9]{1,4} ', line).group(0).strip()
                    name = line[0: line.index(rating)].strip()
                    newr = None
                    data = line[line.index(rating):]
                    data = data.split(';')
                    opponents = data[0].strip().split(' ')[1:]
                    scores = data[1].strip().split(' ')
                    p12 = None
                    rank = None

                    offed = False

                    for d in data:
                        obj = d.strip().split(' ')
                        obj_name = obj[0].strip()
                        itms = obj[1:]
                        
                        if obj_name == 'p12':
                            p12 = itms
                        elif obj_name == 'newr':
                            newr = d.strip().split(' ')[1:]

                        elif obj_name == 'rrank':
                            # important change on Nov 29, 2015 - it was decided to always 
                            # use rcrank instead of rrank since rrank was noticed to have 
                            # the wrong information at times example SOY3 2015

                            # Exception will be when the rcrank field has fewer data items
                            # than the rrank field
                            tmp_rank = itms 
                            if not rank:
                                rank = tmp_rank
                            else :
                                if len(tmp_rank) > len(rank):
                                    rank = tmp_rank

                        elif obj_name == 'rcrank' and rank == None:
                            rank = [0] + itms
                        elif obj_name == 'off':
                            offed = True
                           

                    if rank != None:
                        if opponents and len(rank) > len(opponents) :
                            rank = rank[1:]

                    if not p12:
                        p12 = ['3'] * (len(rank)+1)

                    players.append({'name': name, 'opponents' :opponents,'scores':scores,
                        'p12': p12, 'rank': rank,'newr': newr, 'off': offed, 
                        'old_rating': rating})
                    
                    
            if len(players) < 2:
                print 'a.t file does not contain any data'
                return

            players[0]['scores'] = [0] * len(players[1]['scores'])
            
    def random_results(self, count):
        '''
        Generates random results for simulation purposes.
        '''
        scores = []
        low = 300
        high = 600
        for i in range(count):
            scores.append([random.randrange(low,high),
                random.randrange(low-50,high-50)])

            low -= 1
            high -= 1

        return scores

            
class TshXl(Tsh):
    '''
    Saves processed TSH data into a spreadsheet
    '''
    MAIN_SHEET = 'Standings'

    def __init__(self, filename=None, division_file=None):
        super(TshXl,self).__init__(filename)
    
        if filename:
            self.wb = load_workbook(filename=self.filename)
            sheet = self.wb[TshXl.MAIN_SHEET]
    
            assert sheet['A1'].value == 'Player'
            assert sheet['B1'].value == 'Rating'
            assert sheet['C1'].value == 'Wins'
            assert sheet['D1'].value == 'Spread'

    
    def export_rounds(self, first, last):
        ''' 
        Generates a series of tsh commands.
        These can be piped into tsh to update the division file. For each
        round a set of pairing commands will be generated followed by a 
        series of add score commands.
        '''

        wb = load_workbook(filename=self.filename)
        players = [p[0] for p in self.get_seeded_players()]

        try :
            for i in range(first,last+1) :
                sheet = wb['Round{0}'.format(i)]
                for row in sheet.iter_rows():
                    if row[0].value == 'Bye':
                        pl1 = 0
                    else :
                        pl1 = players.index(row[0].value) +1
                    if row[2].value == 'Bye':
                        pl2 = 0
                    else :
                        pl2 = players.index(row[2].value) +1

                    print 'pair {0} {1} {2}'.format(pl1, pl2, i) 
            print ''

        except KeyError:
            print 'Round data does not exist'
        except IndexError:
            print 'Player not found'
        
        try :
            for i in range(first,last+1) :
                sheet = wb['Round{0}'.format(i)]
                print 'a {0}'.format(i)

                for row in sheet.iter_rows():
                    if not (row[2].value == 'Bye' or row[0].value == 'Bye'):
                        pl1 = players.index(row[0].value)+1
                        pl2 = players.index(row[2].value)+1
                       
                        print pl1,row[1].value, pl2, row[3].value

                print ''
        except KeyError:
            print 'Round data does not exist'


    def save_to_xl(self, filename):
        players = self.players
        wb = Workbook()
        sheet = wb.create_sheet('Standings')
        sheet.append(['Player','Rating','Wins','Margin'])

        rows = []
        rounds = [ [] for i in range(len(players[1]['opponents']))] 

        for idx, player in enumerate(players):
            if player['name'] == 'Bye':
                continue

            row = [player['name'], player['old_rating'],0,0]

            wins = 0
            spread = 0
            
            
            for i,opposite in enumerate(player['opponents']) :

                opponent = players[int(opposite)]
                row.append(opponent['name'])

                try:
                    player_score = player['scores'][i]
                    opponent_score = opponent['scores'][i]

                    if opponent['name'] == 'Bye' :
                        margin = int(player['scores'][i])
                    else :
                        margin = int(player['scores'][i]) - int(opponent['scores'][i])

                    if margin > 0 :
                        win = 1
                    elif margin == 0 :
                        win = 0.5
                    else :
                        win = 0

                    wins += win
                    row.append(win)
                    row.append(margin)

                    spread += margin
                    p = Pair(player['name'],opponent['name'], player_score, opponent_score)
                except IndexError:
                    # this round doesn't have scores yet
                    p = Pair(player['name'],opponent['name'], '','')

                if p not in rounds[i]:
                    rounds[i].append(p)

            row[2] = wins
            row[3] = spread
            rows.append(row)

        for row in sorted(rows, key=lambda x: (x[2],x[3]), reverse = True):
            sheet.append(row)

        for i, rnd in enumerate(rounds):
            sheet = wb.create_sheet('Round{0}'.format(i))
            for p in rnd:
                sheet.append(p.to_array())

        wb.save(filename)
    
    def get_seeded_players(self):
        sheet = self.wb[TshXl.MAIN_SHEET]
        rows = [(row[0].value, row[1].value or 0) for row in sheet.iter_rows(min_row=2)]
        return sorted(rows, key=lambda x: (-x[1],x[0]))

    def create_division(self, players=None):
        
        for row in self.get_seeded_players(): 
            print row[0],' ',row[1]


def simulate(count):
    tsh = TshXl('swiss.xlsx')
    print tsh.random_results(int(count))

def excel_to_division(filename):
    '''
    Generate a tsh division file
    '''
    tsh = TshXl(filename)
    tsh.create_division()

def division_to_excel(division_file, excel_file):
    '''
    Import data from tsh into a spreadsheet
    '''
    tsh = TshXl(division_file)
    tsh.process_data()
    tsh.save_to_xl(excel_file)

def excel_to_tsh_pairs(filename, first, last):
    '''
    Generate tsh pairing commands.
    The pairing is actually done with our code and it's saved in a spreadsheet.
    This data can be exported to tsh via generated tsh pairing commands. The 
    scores are also exported via tsh add scores commands
    '''

    tsh = TshXl(filename)
    tsh.export_rounds(int(first), int(last))


if __name__ == '__main__': #pragma nocover

    argh.dispatch_commands([excel_to_division, division_to_excel, excel_to_tsh_pairs, simulate])

