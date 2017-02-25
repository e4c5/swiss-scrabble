import sys, os
import math
import argh

from argh.decorators import arg

from openpyxl import load_workbook


class Pairing(object):
    '''
    Pairing procedure according to Swiss system rules.
    Adapted from 
     https://github.com/gnomeby/swiss-system-chess-tournament/blob/master/tournaments/pairing.py
    
    '''
        
    def make_it(self):
        if self.next_round == 1:
            self.pair_first_round()
        else:
            self.pair_next_round()
        
        return self.pairs


    def get_detailed_pairings(self):
        pairs = []
        for pair in self.pairs :
            pairs.append(pair)
            pairs.append([pair[1],pair[0]])
        
        return sorted(pairs, reverse=True, key=lambda x: (x[0]['score'], x[0]['spread']))

    def order_players(self, players):
        sorted_players = sorted(players, reverse=True, 
                key=lambda player: (player['score'],player['spread']))
        return sorted_players
    
    def asign_bye(self):
        '''
        Asign a bye to the lowest ranked player who hasn't already been
        asigned one
        '''
        if len(self.players) % 2 == 1:
            sorted_brackets_keys = sorted(self.brackets)
            for group_score in sorted_brackets_keys:
                group = self.brackets[group_score]
                
                for player in reversed(group):
                    if player.has_key('pair') and player['pair']:
                        continue
                    
                    if 'Bye' not in player['opponents'] :
                        player['pair'] = True
                        self.pairs.append([player,self.bye])
                        return;


    def pair_first_round(self):
        
        sorted_players = self.order_players(self.players)
        S1count = len(self.players) / 2
        for index in range(S1count):
            self.pairs.append([sorted_players[index], sorted_players[S1count+index]])
        if len(self.players) % 2 == 1:
            self.pairs.append([sorted_players[S1count*2+1], None])
    
    
    def pair_next_round(self):
        sorted_brackets_keys = sorted(self.brackets, reverse=True)
        highest_group = sorted_brackets_keys[0]
        lowest_group = sorted_brackets_keys[-1:]
        downfloaters = []

        self.asign_bye()

        for group_score in sorted_brackets_keys:
            group = self.brackets[group_score]
            if len(downfloaters) > 0:
                # TODO: B.5, B.6
                group[0:0] = downfloaters
                
                
            downfloaters = []
            for player in group:
                if player.has_key('pair') and player['pair']:
                    continue
                
                opponents = self.find_possible_opponents(player, group)

                # C.1: B.1, B.2
                if len(opponents) == 0:
                    player['downfloater'] = True

                    downfloaters.append(player)
                elif len(opponents) == 1:
                    if player.has_key('downfloater') and player['downfloater']:
                        opponents[0]['upfloater'] = True
                    playerW, playerB = self.return_with_color_preferences(player, opponents[0])
                    self.pairs.append([playerW, playerB])
                    playerW['pair'] = True
                    playerB['pair'] = True
                elif len(opponents) > 1 :
                    sorted_players = self.order_players(opponents)
                    
                    if player.has_key('downfloater') and player['downfloater']:
                       sorted_players[0]['upfloater'] = True
                    playerW, playerB = self.return_with_color_preferences(player, sorted_players[0])
                    self.pairs.append([playerW, playerB])
                    playerW['pair'] = True
                    playerB['pair'] = True
                    pass
                    
            without_opponents = [pl for pl in group if not pl.has_key('pair') or pl['pair'] is False]
            if len(without_opponents) > 2:

                self.pair_group_with_transposition(without_opponents)
                without_opponents = [pl for pl in group if not pl.has_key('pair') or pl['pair'] is False]
                
                if len(without_opponents) == 1:
                    without_opponents[0]['downfloater'] = True
                    downfloaters.append(without_opponents[0]) 
            if len(downfloaters) > 0 and lowest_group == group_score:
                pass
        pass
    
    # D 1.1 Homogenius transposition
    def pair_group_with_transposition(self, group):
        sorted_players = self.order_players(group)
        S1count = len(sorted_players) / 2
        S2count = len(sorted_players) - S1count
        S1 = sorted_players[:S1count]
        S2 = sorted_players[S1count:]

        def transposition(k,n):
            if k == n:
                yield S2
            else:
                for i in range(k, n):
                    if i != k:
                        S2[k], S2[i] = S2[i], S2[k]
                    for x in transposition(k+1, n):
                        yield x
                    if i != k:
                        S2[k], S2[i] = S2[i], S2[k]
            pass
        

        # Check simple pairing
        for S2 in transposition(0, S2count):
            problems = 0
            for index in range(S1count):
                if S2[index] not in self.find_possible_opponents(S1[index], group):
                    problems += 1
                    break
                    
            if problems > 0:
                continue
            
            # Pairing
            for index in range(S1count):
                playerW, playerB = self.return_with_color_preferences(S1[index], S2[index])
                self.pairs.append([playerW, playerB])
                playerW['pair'] = True
                playerB['pair'] = True
                
            return group
                
        return group
    
    def return_with_color_preferences(self, playerA, playerB):
        player1, player2 = self.order_players([playerA, playerB])
        player1_pref = self.get_color_preferences(player1)
        player2_pref = self.get_color_preferences(player2)
        player1_switched_color = self.get_switched_color_for_latest_game(player1)
        player2_switched_color = self.get_switched_color_for_latest_game(player2)
        
        if player1_pref <= -2 or player2_pref >= 2:
            return player1, player2
        elif player1_pref == -1 or player2_pref == 1:
            return player1, player2
        elif player1_pref >= 2 or player2_pref <= -2:
            return player2, player1
        elif player1_pref == 1 or player2_pref == -1:
            return player2, player1
        elif player1_switched_color and player1_switched_color == 'W':
            return player1, player2
        elif player2_switched_color and player2_switched_color == 'W':
            return player2, player1
        
        return player1, player2
    
    def find_player_by_name(self, name):
        for player in self.players:
            if player['name'] == name:
                return player
        return None

    # B.1, B.2    
    def find_possible_opponents(self, current_player, group, skip_color_pref = False):
        rest = []
        idx = group.index(current_player)
        c = len(group) / 2
        if c > idx:
            for player in group[c:]:
                if current_player != player :
                    if not player.has_key('pair') or player['pair'] is False:
                        if player['name'] not in current_player['opponents'] : 
                            rest.append(player)

        if len(rest) == 0:
            for player in group:
                if current_player != player :
                    if not player.has_key('pair') or player['pair'] is False:
                        if player['name'] not in current_player['opponents'] : 
                            rest.append(player)

        if len(rest) == 0:
            return []
        
        # B.2
        # TODO: Top scored players
        color_pref = self.get_color_preferences(current_player)
        if abs(color_pref) >= 2 and skip_color_pref is False:
            for player in rest:
                opponent_color_pref = self.get_color_preferences(player)
                if opponent_color_pref == color_pref:
                    rest.remove(player)
                    continue
                 
        return rest
    
    def get_color_preferences(self, player):
        return 0

    def get_switched_color_for_latest_game(self, player):
        return None
    

class XlPairing(Pairing):
    ''' Pairing from tournament results stored in excel files '''
    HEADER_ROWS = 1

    def __init__(self, filename, next_round=1):

        ''' 
        players - list of dicts(name, rating, spread)
        games - list of dicts(round, player(name), opponent(name||None), player_score, opponent_score, player_color, opponent_color, is_walkover)
        round_number - number of next round
        '''

        self.filename = filename

        self.wb = load_workbook(filename=filename)
        sheet = self.wb['Standings']

        self.players = []
        self.pairs = []
        self.bye = {'name': 'Bye', 'rating': 0, 'score': 0 , 'spread': 0 }
        for pl in sheet.iter_rows(min_row=XlPairing.HEADER_ROWS+1):
            opponents = []
            spread = 0
            wins = 0
            if pl[0].value == None:
                break

            for game in range(1, next_round):
                spread += pl[game*3 + 3].value
                opponents.append(pl[game*3 +1].value)
                wins += pl[game*3 + 2].value
            
            try:
                if pl[0] != 'Bye':
                    self.players.append( 
                         { 'name': pl[0].value,
                           'spread' : spread,
                           'rating' : pl[1].value or 0,
                           'score'  : wins,
                           'opponents' : opponents
                         }
                    )
            except :
                traceback.print_exc() 
                pass
        
        self.next_round = next_round
        
        # A.3
        brackets = {}
        for player in self.players:
            if player['score'] not in brackets:
                brackets[player['score']] = []
            brackets[player['score']].append(player)
        self.brackets = brackets


    def make_it(self):
        pairs = super(XlPairing,self).make_it()
        sheet = self.wb.create_sheet('Round{0}'.format(self.next_round))
        
        for item in sorted(pairs, reverse=True, key=lambda x: (x[0]['score'], x[0]['spread'])):
            sheet.append([item[0]['name'],'', item[1]['name'],''])

        return pairs

    def update_scores(self):
        self.wb = load_workbook(filename=self.filename)
        sheet = self.wb['Standings']

        players = []
        for pl in sheet.iter_rows(min_row=XlPairing.HEADER_ROWS+1,max_col=2):
            players.append(pl[0].value)


        results = self.wb['Round{0}'.format(self.next_round)]
        column = 5 + (self.next_round-1) * 3
        
        print 'Updating results for round ', self.next_round, 'starting from column', column

        for game in results.iter_rows(max_col=5):
            player = players.index(game[0].value) + XlPairing.HEADER_ROWS+1

            if game[2].value == 'Bye':
                opponent = None
            else :
                opponent = players.index(game[2].value) + XlPairing.HEADER_ROWS+1

            spread = game[1].value - game[3].value
            print player, column, 'opponent', opponent

            if spread == 0:
                sheet.cell(row=player,column=column+1,value=0.5)
                sheet.cell(row=opponent,column=column+1,value=0.5)
            
            elif spread > 0:
                sheet.cell(row=player,column=column+1,value=1)
                if opponent:
                    sheet.cell(row=opponent,column=column+1,value=0)
            else :
                sheet.cell(row=player,column=column+1,value=0)
                sheet.cell(row=opponent,column=column+1,value=1)
            
            sheet.cell(row=player,column=column+2,value=spread)
            sheet.cell(row=player,column=column,value=game[2].value)
            
            if opponent:
                sheet.cell(row=opponent,column=column,value=game[0].value)
                sheet.cell(row=opponent,column=column+2,value=-spread)
               


    def save_sheet(self):
        self.wb.save(self.filename)


def update_results(round_number):
    pair = XlPairing('/home/raditha/Downloads/Swiss.xlsx',int(round_number))
    pair.update_scores()
    pair.save_sheet()

def make_pairing(round_number):
    pair = XlPairing('swiss.xlsx',int(round_number))
    pair.make_it()
    pair.save_sheet()


if __name__ == '__main__': #pragma nocover
    argh.dispatch_commands([update_results, make_pairing])

    #pair = Pairing(tournament = Tournament.objects.get(pk=245), next_round = 3)
    #for item in sorted(pair.make_it(), reverse=True, key=lambda x: x[0]['spread']):
    #    print item[0]['name'], item[1]['name']

    #pair = XlPairing('/home/raditha/Downloads/Swiss.xlsx',4)

    #pair.update_scores()
    
    #for item in sorted(pair.make_it(), reverse=True, key=lambda x: x[0]['spread']):
    #    print item[0]['name'], item[1]['name']

#    pair.write_it()

    #pair.write_it()

